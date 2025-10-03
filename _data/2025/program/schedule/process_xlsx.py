#!/usr/bin/env python3
"""
Excel schedule → per day JSON for Jekyll with exact rowspan and colspan,
with trailing empty columns removed and header merges preserved.

This version also accepts an HTTPS URL for --xlsx, for example a published
Google Sheets export link that returns an .xlsx file.
"""

from __future__ import annotations
import argparse, json, os, re, io, urllib.request, urllib.error
from typing import Dict, Tuple, List
from openpyxl import load_workbook

SAFE_PREFIX = "ismar2025_"

def sanitize_basename(s: str) -> str:
    s = str(s).replace("\n", " ")
    s = s.replace("(", "").replace(")", "").replace(".", "").replace(",", "")
    s = re.sub(r"\s+", " ", s).strip().replace(" ", "_")
    return re.sub(r"[^A-Za-z0-9_]", "", s)

def txt(ws, r: int, c: int) -> str:
    v = ws.cell(r, c).value
    return "" if v is None else str(v).strip()

def build_merge_maps(ws):
    """For every cell, store its merge anchor and the rectangle of that anchor."""
    anchor_of = {}
    rect_of = {}

    R, C = ws.max_row, ws.max_column
    # Initialize each cell to its own anchor
    for r in range(1, R + 1):
        for c in range(1, C + 1):
            anchor_of[(r, c)] = (r, c)
            rect_of[(r, c)] = (r, c, r, c)

    # Fill in merged ranges
    for mr in ws.merged_cells.ranges:
        r1, r2 = mr.min_row, mr.max_row
        c1, c2 = mr.min_col, mr.max_col
        ar, ac = r1, c1
        rect_of[(ar, ac)] = (r1, c1, r2, c2)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                anchor_of[(r, c)] = (ar, ac)

    return anchor_of, rect_of

def find_header_rows(ws) -> List[Tuple[int,int]]:
    """Find rows where a cell equals 'Time'."""
    out = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if txt(ws, r, c).lower() == "time":
                out.append((r, c))
                break
    if not out:
        raise RuntimeError("No 'Time' header rows found.")
    return out

def left_label(ws, r: int, upto_col: int) -> str:
    c = upto_col - 1
    while c >= 1:
        t = txt(ws, r, c)
        if t != "":
            return t
        c -= 1
    return f"Day starting row {r}"

def last_used_col_for_block(ws, anchor_of, rect_of, start_row: int, end_row: int, time_col: int) -> int:
    """
    Determine the last column index needed for this day.
    """
    last_c = time_col
    R, C = ws.max_row, ws.max_column

    # Header contributions
    for c in range(time_col, C + 1):
        if txt(ws, start_row, c) != "":
            last_c = max(last_c, c)

    # Body and merged ranges
    for r in range(start_row + 1, end_row + 1):
        for c in range(time_col, C + 1):
            if txt(ws, r, c) != "":
                last_c = max(last_c, c)
            ar, ac = anchor_of[(r, c)]
            r1, c1, r2, c2 = rect_of[(ar, ac)]
            if r1 <= r <= r2 and c2 > last_c:
                last_c = c2

    return last_c

def trim_trailing_empty_cols(ws, time_col: int, last_col: int, start_row: int, end_row: int) -> int:
    """Remove trailing columns that are completely empty in both header and body."""
    c = last_col
    while c > time_col:
        all_empty = True
        for r in range(start_row, end_row + 1):
            if txt(ws, r, c) != "":
                all_empty = False
                break
        if all_empty:
            c -= 1
        else:
            break
    return c

def is_url(s: str) -> bool:
    return s.startswith("http://") or s.startswith("https://")

def fetch_xlsx_bytes(url: str, timeout: int = 30) -> bytes:
    """
    Retrieve an .xlsx file over HTTPS and return its raw bytes.
    Follows redirects. Uses a simple User Agent to avoid blocked requests.
    """
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "Python-urllib/3.x xlsx-fetch"}
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = resp.read()
            if not data:
                raise RuntimeError("Downloaded file is empty.")
            return data
    except urllib.error.URLError as e:
        raise RuntimeError(f"Failed to fetch xlsx from URL, reason: {e}") from e

def load_first_worksheet(xlsx_source: str):
    """
    Load the first worksheet from either a local path or an HTTPS URL.
    """
    if is_url(xlsx_source):
        raw = fetch_xlsx_bytes(xlsx_source)
        bio = io.BytesIO(raw)
        wb = load_workbook(bio, data_only=True)
    else:
        wb = load_workbook(xlsx_source, data_only=True)
    return wb.worksheets[0]

def export_json(xlsx: str, out_dir: str):
    os.makedirs(out_dir, exist_ok=True)
    ws = load_first_worksheet(xlsx)
    anchor_of, rect_of = build_merge_maps(ws)
    header_rows = find_header_rows(ws)

    blocks = []
    for i, (hr, tc) in enumerate(header_rows):
        start = hr
        end = header_rows[i + 1][0] - 1 if i + 1 < len(header_rows) else ws.max_row
        blocks.append((start, end, tc))

    index = {"days": []}

    for start, end, time_col in blocks:
        day_label = left_label(ws, start, time_col)
        last_col = last_used_col_for_block(ws, anchor_of, rect_of, start, end, time_col)

        # Trim trailing completely empty columns
        last_col = trim_trailing_empty_cols(ws, time_col, last_col, start, end)

        # Headers with merge detection
        headers = []
        for c in range(time_col, last_col + 1):
            ar, ac = anchor_of[(start, c)]
            r1, c1, r2, c2 = rect_of[(ar, ac)]
            render = (r1 == start and c == c1)
            colspan = min(c2, last_col) - c1 + 1
            headers.append({
                "text": txt(ws, r1, c1) or ("Time" if c == time_col else ""),
                "render": render,
                "colspan": colspan
            })

        # Body rows
        rows_out = []
        for r in range(start + 1, end + 1):
            if all(txt(ws, r, c) == "" for c in range(time_col, last_col + 1)):
                continue
            time_txt = txt(ws, r, time_col)
            cells = []
            for c in range(time_col + 1, last_col + 1):
                ar, ac = anchor_of[(r, c)]
                r1, c1, r2, c2 = rect_of[(ar, ac)]
                render = (r == r1 and c == c1)
                if render:
                    c2_clipped = min(c2, last_col)
                    text = txt(ws, r1, c1)
                    rowspan = r2 - r1 + 1
                    colspan = c2_clipped - c1 + 1
                    cells.append({
                        "text": text,
                        "rowspan": rowspan,
                        "colspan": colspan,
                        "render": True
                    })
                else:
                    cells.append({"render": False})
            rows_out.append({"time": time_txt, "cells": cells})

        base = SAFE_PREFIX + sanitize_basename(day_label)
        with open(os.path.join(out_dir, f"{base}.json"), "w", encoding="utf-8") as f:
            json.dump({"date": day_label, "headers": headers, "rows": rows_out}, f, ensure_ascii=False, indent=2)

        index["days"].append({"label": day_label, "file": f"{base}.json"})

    with open(os.path.join(out_dir, "schedule_index.json"), "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)

def main():
    ap = argparse.ArgumentParser(description="Excel to JSON with exact merges and continuous per day grids. Accepts a local .xlsx path or an HTTPS URL.")
    ap.add_argument("--xlsx", required=False, help="Path to the Excel file, or an HTTPS link to an .xlsx export", default="https://docs.google.com/spreadsheets/d/e/2PACX-1vQrmgHgmh-sZFvt8d_Wz2Ma69b1l4xsuwLbSiNrNKH2BgJmTFTEPLN7fw6HqHWjLKJK7i48kBKjE-K9/pub?output=xlsx")
    ap.add_argument("--out", required=False, default="./json", help="Target folder")
    ap.add_argument("--timeout", type=int, default=30, help="Network timeout in seconds when fetching a URL")
    args = ap.parse_args()

    # If a URL is provided, validate early by attempting a small fetch, then proceed
    if is_url(args.xlsx):
        # Pre fetch to fail fast with a clearer message, then export_json will refetch only once if needed
        try:
            _ = fetch_xlsx_bytes(args.xlsx, timeout=args.timeout)
        except Exception as e:
            raise SystemExit(f"Error fetching --xlsx URL: {e}")

    export_json(args.xlsx, args.out)
    print(f"Wrote JSON to {args.out}")

if __name__ == "__main__":
    main()
